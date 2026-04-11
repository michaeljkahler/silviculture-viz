"""
generate_xlsx_template.py — Erstellt waldbau-state_VORLAGE.xlsx

Eine benutzerfreundliche Excel-Vorlage mit:
- Eingabe-Sheet: Dropdowns und Validierungen für alle Parameter
- CSV-Export-Sheet: Formel-basierte CSV-Ausgabe (zum Kopieren oder Save-as-CSV)
- Anleitung-Sheet: Dokumentation in Deutsch

Ausführung: python tools/generate_xlsx_template.py
Erzeugt: waldbau-state_VORLAGE.xlsx im Repo-Root
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# === Konstanten ===
SPECIES = [
    ('Bu',   'Buche',       'Fagus sylvatica',     65),
    ('Ta',   'Weisstanne',  'Abies alba',          15),
    ('Fi',   'Fichte',      'Picea abies',         10),
    ('BAh',  'Ahorn',       'Acer ssp.',           10),
    ('Ei',   'Eiche',       'Quercus ssp.',         0),
    ('WFoe', 'Waldfoehre',  'Pinus sylvestris',     0),
    ('Lae',  'Laerche',     'Larix decidua',        0),
    ('Es',   'Esche',       'Fraxinus excelsior',   0),
    ('Dou',  'Douglasie',   'Pseudotsuga menz.',    0),
    ('Li',   'Linde',       'Tilia ssp.',           0),
    ('Bi',   'Birke',       'Betula pendula',       0),
]

MIX_PATTERNS = ['einzel', 'trupp', 'gruppe', 'horst']
MODES = ['gleich', 'ungleich']
BOOL_VALUES = ['true', 'false']

# Zentrale Defaults pro Baumart (h, cw, kl, ka, bhd)
SPECIES_DEFAULTS = {
    'Bu':   (35, 12, 20, 15, 50),
    'Ta':   (40,  8, 16, 24, 55),
    'Fi':   (38,  6, 13, 25, 45),
    'BAh':  (30, 10, 18, 12, 40),
    'Ei':   (30, 14, 18, 12, 55),
    'WFoe': (28,  8,  9, 19, 40),
    'Lae':  (35,  7, 12, 23, 40),
    'Es':   (32, 10, 18, 14, 45),
    'Dou':  (42,  7, 14, 28, 55),
    'Li':   (32, 11, 18, 14, 50),
    'Bi':   (25,  8, 14, 11, 35),
}

# === Styles ===
FONT_BASE = Font(name='Arial', size=10)
FONT_HEADER = Font(name='Arial', size=12, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Arial', size=14, bold=True, color='2a6a3a')
FONT_SUBTITLE = Font(name='Arial', size=11, bold=True, color='2a6a3a')
FONT_HELP = Font(name='Arial', size=9, italic=True, color='666666')
FONT_INPUT = Font(name='Arial', size=10, color='0000FF')

FILL_HEADER = PatternFill('solid', start_color='2a6a3a')
FILL_SUBHEADER = PatternFill('solid', start_color='d4e8d4')
FILL_INPUT = PatternFill('solid', start_color='fffae6')
FILL_ALT = PatternFill('solid', start_color='f7f6f2')
FILL_INFO = PatternFill('solid', start_color='e8f0e8')

BORDER_THIN = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)


def style_header_cell(cell):
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN


def style_input_cell(cell):
    cell.font = FONT_INPUT
    cell.fill = FILL_INPUT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN


def style_label_cell(cell):
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = BORDER_THIN


def style_help_cell(cell):
    cell.font = FONT_HELP
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def main():
    wb = Workbook()

    # ===========================================================
    # Sheet 1: Anleitung
    # ===========================================================
    ws_help = wb.active
    ws_help.title = 'Anleitung'

    ws_help.column_dimensions['A'].width = 100

    ws_help['A1'] = 'waldbau-viz Arbeitsstand-Vorlage'
    ws_help['A1'].font = Font(name='Arial', size=18, bold=True, color='2a6a3a')

    ws_help['A2'] = 'Schemaversion: 1'
    ws_help['A2'].font = FONT_HELP

    instructions = [
        '',
        'ANLEITUNG',
        '',
        '1. Wechsle zum Tab "Eingabe" und fülle die Felder aus.',
        '   • Dropdown-Felder zeigen einen kleinen Pfeil — wähle einen erlaubten Wert.',
        '   • Zahlenfelder werden automatisch validiert (z.B. Prozent muss zwischen 0 und 100 sein).',
        '   • Gelb hinterlegte Zellen sind Eingabefelder — ändere KEINE anderen Zellen.',
        '',
        '2. Wechsle zum Tab "CSV-Export".',
        '   • Hier wird automatisch das CSV-Format aus deinen Eingaben aufgebaut.',
        '   • Markiere alle Zellen in Spalte A (Strg+A oder Klick auf "A" und dann Strg+Shift+End).',
        '   • Kopiere mit Strg+C.',
        '',
        '3. Öffne einen Texteditor (Notepad, Notepad++, VS Code, ...).',
        '   • Füge mit Strg+V ein.',
        '   • Speichere als "waldbau-state.csv" (Encoding: UTF-8).',
        '',
        '4. Im waldbau-viz Tool: Klick auf "📥 CSV Laden" und wähle deine Datei.',
        '',
        'ALTERNATIVE: Direkt als CSV speichern',
        '',
        '   • Wechsle zum "CSV-Export"-Tab',
        '   • Datei → Speichern unter → Format: "CSV (Trennzeichen-getrennt) (.csv)"',
        '   • Achtung: Excel speichert je nach Region ; oder , als Trenner.',
        '     Falls Probleme: Verwende den Copy-Paste-Workflow oben.',
        '',
        'HINWEISE',
        '',
        '• Werte für Höhe, Kronenbreite etc. (h, cw, kl, ka, bhd) sind nur im manuellen Modus relevant.',
        '  Bei timelineOn=true werden sie aus dem Alter berechnet (FBB/SiWaWa-Kurven).',
        '',
        '• Die Summe aller pct-Werte aktiver Baumarten sollte 100 ergeben (wird beim Import nicht erzwungen).',
        '',
        '• Dienerbäume: Mehrere Arten mit Semikolon trennen (z.B. "Ta;Bu;BAh").',
        '  Nur Schattenbaumarten erlaubt: Ta, Bu, BAh, Fi, Li, Es, Dou.',
        '',
        '• mixOverflow: Zweite Mischungsform falls die Primärform vollständig auf der Fläche untergebracht ist.',
        '  Leer lassen wenn nicht gewünscht.',
        '',
        'BAUMARTEN-CODES',
        '',
        '   Bu   = Buche       (Fagus sylvatica)',
        '   Ta   = Weisstanne  (Abies alba)',
        '   Fi   = Fichte      (Picea abies)',
        '   BAh  = Ahorn       (Acer ssp.)',
        '   Ei   = Eiche       (Quercus ssp.)',
        '   WFoe = Waldföhre   (Pinus sylvestris)',
        '   Lae  = Lärche      (Larix decidua)',
        '   Es   = Esche       (Fraxinus excelsior)',
        '   Dou  = Douglasie   (Pseudotsuga menziesii)',
        '   Li   = Linde       (Tilia ssp.)',
        '   Bi   = Birke       (Betula pendula)',
        '',
        'MISCHUNGSFORMEN',
        '',
        '   einzel = Einzelmischung (zufällige Verteilung)',
        '   trupp  = Trupp (~12m Cluster, 5–15 Bäume)',
        '   gruppe = Gruppe (~18m Cluster, 15–30 Bäume)',
        '   horst  = Horst (~30m Cluster, 30–50 Bäume)',
        '',
        'SCHICHTEN',
        '',
        '   layerO = Oberschicht (Hauptbestand)',
        '   layerM = Mittelschicht',
        '   layerU = Unterschicht',
        '   layerV = Verjüngung',
        '',
    ]

    for i, line in enumerate(instructions, start=3):
        cell = ws_help.cell(row=i, column=1, value=line)
        if line in ('ANLEITUNG', 'ALTERNATIVE: Direkt als CSV speichern',
                    'HINWEISE', 'BAUMARTEN-CODES', 'MISCHUNGSFORMEN', 'SCHICHTEN'):
            cell.font = Font(name='Arial', size=12, bold=True, color='2a6a3a')
        else:
            cell.font = FONT_BASE
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # ===========================================================
    # Sheet 2: Eingabe
    # ===========================================================
    ws_in = wb.create_sheet('Eingabe')

    ws_in.column_dimensions['A'].width = 22
    ws_in.column_dimensions['B'].width = 15
    ws_in.column_dimensions['C'].width = 60

    # === Titel ===
    ws_in.merge_cells('A1:C1')
    ws_in['A1'] = 'waldbau-viz Arbeitsstand — Eingabe'
    ws_in['A1'].font = FONT_TITLE
    ws_in['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_in.row_dimensions[1].height = 28

    # === Globale Einstellungen ===
    ws_in.merge_cells('A3:C3')
    ws_in['A3'] = 'GLOBALE EINSTELLUNGEN'
    ws_in['A3'].font = FONT_HEADER
    ws_in['A3'].fill = FILL_HEADER
    ws_in['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws_in.row_dimensions[3].height = 22

    globals_data = [
        # (label, default value, help text, validation type)
        ('mode',              'gleich', 'Waldtyp: gleich = gleichförmiger Hochwald, ungleich = Plenterwald', ('list', MODES)),
        ('density',           220,      'Stämme pro Hektar (60–600)', ('number', 60, 600)),
        ('transectY',         50,       'Y-Position des Transekts in Metern (5–95)', ('number', 5, 95)),
        ('transectWidth',     10,       'Breite des Transekts in Metern (5–15)', ('number', 5, 15)),
        ('seed',              42,       'Zufallszahl für reproduzierbare Generierung (1–9999)', ('number', 1, 9999)),
        ('age',               0,        'Bestandesalter in Jahren — nur relevant wenn timelineOn=true (0–140)', ('number', 0, 140)),
        ('timelineOn',        'false',  'Zeitschiene aktiv? (Bei true werden h/cw/kl/ka/bhd aus age berechnet)', ('list', BOOL_VALUES)),
        ('showVerjuengung',   'true',   'Naturverjüngung anzeigen?', ('list', BOOL_VALUES)),
        ('showZBaum',         'false',  'Z-Bäume hervorheben?', ('list', BOOL_VALUES)),
        ('showTotholz',       'false',  'Totholz (stehend + liegend) anzeigen?', ('list', BOOL_VALUES)),
        ('showHabitat',       'false',  'Habitatbäume markieren?', ('list', BOOL_VALUES)),
        ('nHabitat',          5,        'Anzahl Habitatbäume pro ha (3–5)', ('number', 0, 20)),
        ('dienerRadius',      10,       'Breite des Dienerbaum-Mantels in Metern', ('number', 5, 20)),
        ('mixPatternDefault', 'einzel', 'Standard-Mischungsform', ('list', MIX_PATTERNS)),
    ]

    # Header row
    for col, header in enumerate(['Parameter', 'Wert', 'Beschreibung'], start=1):
        cell = ws_in.cell(row=4, column=col, value=header)
        style_header_cell(cell)
    ws_in.row_dimensions[4].height = 20

    # Data validations
    dv_mode = DataValidation(type='list', formula1='"' + ','.join(MODES) + '"', allow_blank=False)
    dv_bool = DataValidation(type='list', formula1='"' + ','.join(BOOL_VALUES) + '"', allow_blank=False)
    dv_mix = DataValidation(type='list', formula1='"' + ','.join(MIX_PATTERNS) + '"', allow_blank=False)
    dv_mix_or_empty = DataValidation(type='list', formula1='",einzel,trupp,gruppe,horst"', allow_blank=True)

    ws_in.add_data_validation(dv_mode)
    ws_in.add_data_validation(dv_bool)
    ws_in.add_data_validation(dv_mix)
    ws_in.add_data_validation(dv_mix_or_empty)

    # Number validations (per range)
    num_validations = {}

    def get_num_dv(min_val, max_val):
        key = (min_val, max_val)
        if key not in num_validations:
            dv = DataValidation(
                type='whole',
                operator='between',
                formula1=str(min_val),
                formula2=str(max_val),
                allow_blank=False,
                error='Wert muss zwischen {} und {} liegen.'.format(min_val, max_val),
                errorTitle='Ungültiger Wert',
            )
            num_validations[key] = dv
            ws_in.add_data_validation(dv)
        return num_validations[key]

    # Fill globals
    for i, (label, default, helptext, validation) in enumerate(globals_data, start=5):
        # Label
        cell = ws_in.cell(row=i, column=1, value=label)
        style_label_cell(cell)
        cell.fill = FILL_SUBHEADER

        # Value
        val_cell = ws_in.cell(row=i, column=2, value=default)
        style_input_cell(val_cell)

        # Apply validation
        if validation[0] == 'list':
            opts = validation[1]
            if opts == MODES:
                dv_mode.add(val_cell)
            elif opts == BOOL_VALUES:
                dv_bool.add(val_cell)
            elif opts == MIX_PATTERNS:
                dv_mix.add(val_cell)
        elif validation[0] == 'number':
            dv_num = get_num_dv(validation[1], validation[2])
            dv_num.add(val_cell)

        # Help
        help_cell = ws_in.cell(row=i, column=3, value=helptext)
        style_help_cell(help_cell)
        help_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # === Baumarten-Sektion ===
    species_start_row = 5 + len(globals_data) + 2  # 2 Leerzeilen Abstand

    ws_in.merge_cells(f'A{species_start_row}:P{species_start_row}')
    title_cell = ws_in.cell(row=species_start_row, column=1, value='BAUMARTEN-EINSTELLUNGEN')
    title_cell.font = FONT_HEADER
    title_cell.fill = FILL_HEADER
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_in.row_dimensions[species_start_row].height = 22

    # Erweiterte Spaltenbreiten für Baumarten-Tabelle
    # Format: A=Code, B=Name, C=on, D=pct, E=h, F=cw, G=kl, H=ka, I=bhd,
    #         J=layerO, K=layerM, L=layerU, M=layerV, N=habitat, O=mixPattern, P=mixOverflow, Q=diener
    species_widths = {
        'A': 8,   # Code
        'B': 14,  # Name
        'C': 8,   # on
        'D': 7,   # pct
        'E': 6,   # h
        'F': 6,   # cw
        'G': 6,   # kl
        'H': 6,   # ka
        'I': 6,   # bhd
        'J': 8,   # layerO
        'K': 8,   # layerM
        'L': 8,   # layerU
        'M': 8,   # layerV
        'N': 9,   # habitat
        'O': 12,  # mixPattern
        'P': 12,  # mixOverflow
        'Q': 22,  # diener
    }
    for col_letter, width in species_widths.items():
        if ws_in.column_dimensions[col_letter].width < width:
            ws_in.column_dimensions[col_letter].width = width

    # Header row for species
    species_header_row = species_start_row + 1
    headers = ['Code', 'Name', 'on', 'pct', 'h', 'cw', 'kl', 'ka', 'bhd',
               'layerO', 'layerM', 'layerU', 'layerV', 'habitat',
               'mixPattern', 'mixOverflow', 'diener']
    for col, header in enumerate(headers, start=1):
        cell = ws_in.cell(row=species_header_row, column=col, value=header)
        style_header_cell(cell)

    # Comments für Header
    header_comments = {
        3: 'Baumart aktiv? (true/false)',
        4: 'Mischungsanteil in % (0-100)',
        5: 'Oberhöhe in Metern',
        6: 'Kronenbreite in Metern',
        7: 'Kronenlänge in Metern',
        8: 'Kronenansatz in Metern',
        9: 'Brusthöhendurchmesser in cm',
        10: 'Oberschicht (Hauptbestand)',
        11: 'Mittelschicht',
        12: 'Unterschicht',
        13: 'Verjüngung (Naturverjüngung)',
        14: 'Habitatbaum-Kandidat (grösste Bäume werden geschützt)',
        15: 'Mischungsform: einzel/trupp/gruppe/horst',
        16: 'Zweite Mischungsform falls Primärform voll',
        17: 'Schattendiener (Liste mit Semikolon, z.B. "Ta;Bu;BAh")',
    }
    for col, comment_text in header_comments.items():
        cell = ws_in.cell(row=species_header_row, column=col)
        cell.comment = Comment(comment_text, 'Vorlage')

    ws_in.row_dimensions[species_header_row].height = 22

    # Validation für Baumarten-Spalten
    dv_bool_sp = DataValidation(type='list', formula1='"' + ','.join(BOOL_VALUES) + '"', allow_blank=False)
    dv_mix_sp = DataValidation(type='list', formula1='"' + ','.join(MIX_PATTERNS) + '"', allow_blank=False)
    dv_mix_overflow = DataValidation(type='list', formula1='",einzel,trupp,gruppe,horst"', allow_blank=True)
    dv_pct = DataValidation(type='whole', operator='between', formula1='0', formula2='100', allow_blank=False,
                            error='Prozent muss zwischen 0 und 100 liegen.', errorTitle='Ungültiger Wert')
    dv_pos_num = DataValidation(type='decimal', operator='between', formula1='0', formula2='200', allow_blank=False)
    ws_in.add_data_validation(dv_bool_sp)
    ws_in.add_data_validation(dv_mix_sp)
    ws_in.add_data_validation(dv_mix_overflow)
    ws_in.add_data_validation(dv_pct)
    ws_in.add_data_validation(dv_pos_num)

    # Species rows
    species_first_row = species_header_row + 1
    for sp_idx, (code, name, lat, default_pct) in enumerate(SPECIES):
        row = species_first_row + sp_idx
        h, cw, kl, ka, bhd = SPECIES_DEFAULTS[code]

        on_default = 'true' if default_pct > 0 else 'false'

        # Code (locked label)
        c_code = ws_in.cell(row=row, column=1, value=code)
        c_code.font = Font(name='Arial', size=10, bold=True)
        c_code.fill = FILL_SUBHEADER
        c_code.border = BORDER_THIN
        c_code.alignment = Alignment(horizontal='center', vertical='center')

        # Name
        c_name = ws_in.cell(row=row, column=2, value=name)
        c_name.font = FONT_BASE
        c_name.fill = FILL_SUBHEADER
        c_name.border = BORDER_THIN

        # Inputs (gelb)
        cells = [
            (3, on_default, 'bool'),
            (4, default_pct, 'pct'),
            (5, h, 'pos_num'),
            (6, cw, 'pos_num'),
            (7, kl, 'pos_num'),
            (8, ka, 'pos_num'),
            (9, bhd, 'pos_num'),
            (10, 'true', 'bool'),
            (11, 'false', 'bool'),
            (12, 'false', 'bool'),
            (13, 'false', 'bool'),
            (14, 'false', 'bool'),
            (15, 'einzel', 'mix'),
            (16, '', 'mix_overflow'),
            (17, '', 'text'),
        ]
        for col, val, vtype in cells:
            c = ws_in.cell(row=row, column=col, value=val)
            style_input_cell(c)
            if vtype == 'bool':
                dv_bool_sp.add(c)
            elif vtype == 'pct':
                dv_pct.add(c)
            elif vtype == 'pos_num':
                dv_pos_num.add(c)
            elif vtype == 'mix':
                dv_mix_sp.add(c)
            elif vtype == 'mix_overflow':
                dv_mix_overflow.add(c)

    ws_in.freeze_panes = 'C{}'.format(species_first_row)

    # ===========================================================
    # Sheet 3: CSV-Export (formel-basiert)
    # ===========================================================
    ws_csv = wb.create_sheet('CSV-Export')
    ws_csv.column_dimensions['A'].width = 120
    ws_csv['A1'] = '# Diese Spalte enthält den fertigen CSV-Inhalt — markieren, kopieren, in Texteditor einfügen, als .csv speichern.'
    ws_csv['A1'].font = FONT_HELP

    # Helper: build cell reference for Eingabe sheet
    def in_ref(row, col):
        return f"Eingabe!{get_column_letter(col)}{row}"

    csv_lines = []

    # Header / Metadata (literal text — keine Formel)
    csv_lines.append('# waldbau-viz Arbeitsstand (CSV Format)')
    csv_lines.append('# Schemaversion: 1')
    csv_lines.append('# Erstellt mit waldbau-state_VORLAGE.xlsx')
    csv_lines.append('#')
    csv_lines.append('')
    csv_lines.append('##GLOBALS')
    csv_lines.append('key,value')

    # Globals (rows 5..18 in Eingabe = labels and B-column = values)
    for i, (label, _default, _help, _val) in enumerate(globals_data):
        row = 5 + i
        csv_lines.append(f'="{label},"&{in_ref(row, 2)}')

    csv_lines.append('')
    csv_lines.append('##SPECIES')
    csv_lines.append('species,on,pct,h,cw,kl,ka,bhd,layerO,layerM,layerU,layerV,habitat,mixPattern,mixOverflow,diener')

    # Species rows
    for sp_idx, (code, _name, _lat, _pct) in enumerate(SPECIES):
        row = species_first_row + sp_idx
        # Build formula: =A&","&C&","&D&","&E&","&...
        parts = [f'{in_ref(row, 1)}']  # code
        for col in range(3, 18):  # columns C..Q (on..diener)
            parts.append(f'{in_ref(row, col)}')
        formula = '=' + '&","&'.join(parts)
        csv_lines.append(formula)

    csv_lines.append('')

    # Write csv lines into A column from row 3
    for i, line in enumerate(csv_lines, start=3):
        cell = ws_csv.cell(row=i, column=1, value=line)
        cell.font = Font(name='Consolas', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws_csv['A2'] = '# Tipp: Klicke auf den Spaltenkopf "A", drücke Strg+C, öffne einen Texteditor, drücke Strg+V, speichere als .csv'
    ws_csv['A2'].font = FONT_HELP

    # ===========================================================
    # Save
    # ===========================================================
    out_path = 'waldbau-state_VORLAGE.xlsx'
    wb.save(out_path)
    print(f'OK: {out_path} erstellt')
    print(f'  - {len(globals_data)} globale Parameter')
    print(f'  - {len(SPECIES)} Baumarten')
    print(f'  - {len(csv_lines)} CSV-Zeilen')


if __name__ == '__main__':
    main()
