"""
test_xlsx_template.py — Verifiziert dass waldbau-state_VORLAGE.xlsx
korrekt funktioniert.

Simuliert Excels Formel-Auswertung in Python und füttert das Ergebnis
durch die csvToState() JS-Funktion (via Node.js).

Ausführung: python tests/test_xlsx_template.py
"""

import os
import re
import subprocess
import sys
from openpyxl import load_workbook


def evaluate_formula(formula, get_cell):
    """
    Vereinfachte Auswertung einer Excel-Formel der Form
    ="text"&Eingabe!B5  oder  =Eingabe!A23&","&Eingabe!C23&...
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    expr = formula[1:]  # remove leading =

    # Tokenize: split by & at top level
    parts = []
    current = ''
    depth = 0
    in_quote = False
    for c in expr:
        if c == '"' and not in_quote:
            in_quote = True
            current += c
        elif c == '"' and in_quote:
            in_quote = False
            current += c
        elif c == '&' and not in_quote and depth == 0:
            parts.append(current.strip())
            current = ''
        else:
            current += c
    if current.strip():
        parts.append(current.strip())

    result = ''
    for p in parts:
        if p.startswith('"') and p.endswith('"'):
            result += p[1:-1]
        elif '!' in p:
            # Sheet!Cell reference
            sheet_name, ref = p.split('!')
            val = get_cell(sheet_name, ref)
            result += str(val) if val is not None else ''
        else:
            result += p
    return result


def cell_ref_to_indices(ref):
    """A1 -> (1,1), B5 -> (5,2)"""
    m = re.match(r'^([A-Z]+)(\d+)$', ref)
    if not m:
        return None
    col_letters, row_str = m.groups()
    col = 0
    for c in col_letters:
        col = col * 26 + (ord(c) - ord('A') + 1)
    return int(row_str), col


def main():
    xlsx_path = 'waldbau-state_VORLAGE.xlsx'
    if not os.path.exists(xlsx_path):
        print(f'FEHLER: {xlsx_path} nicht gefunden. Erst tools/generate_xlsx_template.py ausführen.')
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=False)

    def get_cell(sheet_name, ref):
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        idx = cell_ref_to_indices(ref)
        if not idx:
            return None
        return ws.cell(row=idx[0], column=idx[1]).value

    # Evaluate CSV-Export sheet
    ws_csv = wb['CSV-Export']
    csv_lines = []
    for r in range(3, ws_csv.max_row + 1):
        val = ws_csv.cell(row=r, column=1).value
        if val is None:
            csv_lines.append('')
            continue
        if isinstance(val, str) and val.startswith('='):
            evaluated = evaluate_formula(val, get_cell)
            csv_lines.append(evaluated)
        else:
            csv_lines.append(str(val))

    csv_text = '\n'.join(csv_lines)

    print('=== Generated CSV (first 30 lines) ===')
    for i, line in enumerate(csv_lines[:30]):
        print(f'  {i+1:2d}: {line}')
    print()

    # Save to temp file and feed through csvToState via Node.js
    with open('tests/_xlsx_generated.csv', 'w', encoding='utf-8', newline='\n') as f:
        f.write(csv_text)

    # Test via Node.js
    test_script = r'''
const fs = require('fs');
const path = require('path');
const vm = require('vm');

// DOM stubs (minimal)
global.document = {
  getElementById: () => ({ value: '42', style: {}, className: '', textContent: '', querySelector: () => null, querySelectorAll: () => [], appendChild() {}, setAttribute() {} }),
  querySelector: () => null, querySelectorAll: () => [],
  createElement: () => ({ getContext: () => ({}), setAttribute() {}, appendChild() {}, style: {}, click() {}, toDataURL: () => '', cloneNode() { return this; } }),
  createElementNS: () => ({ setAttribute() {}, appendChild() {}, style: {} }),
  body: { appendChild() {}, removeChild() {} },
  addEventListener: () => {},
};
global.window = { addEventListener: () => {}, innerWidth: 1200, innerHeight: 800, devicePixelRatio: 1, requestAnimationFrame: (fn) => fn(), getComputedStyle: () => ({}) };
global.Image = class { constructor() { this.onload = null; } };
global.URL = { createObjectURL: () => 'blob:', revokeObjectURL: () => {} };
global.Blob = class { constructor(p) { this.p = p; this._content = p.join(''); } };
global.XMLSerializer = class { serializeToString() { return ''; } };
global.alert = () => {};
global.requestAnimationFrame = (fn) => fn();
global.navigator = { userAgent: 'node' };
global.HTMLCanvasElement = class {};
global.HTMLElement = class {};
global.DOMParser = class { parseFromString() { return { documentElement: { outerHTML: '' } }; } };
global.FileReader = class {};

const html = fs.readFileSync('index.html', 'utf8');
const matches = [...html.matchAll(/<script>([\s\S]*?)<\/script>/g)];
try { vm.runInThisContext(matches[1][1], { filename: 'index.html' }); } catch(e) { console.error('Init:', e.message); }

const csv = fs.readFileSync('tests/_xlsx_generated.csv', 'utf8');
const data = csvToState(csv);

console.log('=== Geparste Daten ===');
console.log('mode:', data.mode);
console.log('density:', data.density);
console.log('seed:', data.seed);
console.log('timelineOn:', data.timelineOn);
console.log('Anzahl Arten:', Object.keys(data.sp || {}).length);
if (data.sp && data.sp.Bu) {
  console.log('Bu.on:', data.sp.Bu.on);
  console.log('Bu.pct:', data.sp.Bu.pct);
  console.log('Bu.h:', data.sp.Bu.h);
}

// Apply
applyLoadedState(data);
console.log('\n=== State nach Apply ===');
console.log('S.mode:', S.mode);
console.log('S.density:', S.density);
console.log('S.sp.Bu.pct:', S.sp.Bu.pct);
console.log('S.sp.Ta.pct:', S.sp.Ta.pct);
console.log('S.sp.Fi.pct:', S.sp.Fi.pct);
console.log('S.sp.BAh.pct:', S.sp.BAh.pct);
console.log('Sum:', S.sp.Bu.pct + S.sp.Ta.pct + S.sp.Fi.pct + S.sp.BAh.pct);

// Validate critical fields
const errors = [];
if (data.mode !== 'gleich') errors.push('mode != gleich');
if (data.density !== 220) errors.push('density != 220');
if (data.seed !== 42) errors.push('seed != 42');
if (!data.sp || !data.sp.Bu || data.sp.Bu.pct !== 65) errors.push('Bu.pct != 65');
if (!data.sp || !data.sp.Bi) errors.push('Bi missing');

if (errors.length === 0) {
  console.log('\nALLE PRÜFUNGEN BESTANDEN');
  process.exit(0);
} else {
  console.log('\nFEHLER:', errors.join(', '));
  process.exit(1);
}
'''
    with open('tests/_xlsx_validator.js', 'w', encoding='utf-8') as f:
        f.write(test_script)

    result = subprocess.run(
        ['node', 'tests/_xlsx_validator.js'],
        capture_output=True,
        text=True,
        encoding='utf-8',
        env={**os.environ, 'PATH': r'C:\Program Files\nodejs;' + os.environ.get('PATH', '')}
    )
    # Encoding-safe print
    sys.stdout.buffer.write(result.stdout.encode('utf-8'))
    if result.stderr:
        sys.stdout.buffer.write(b'\nSTDERR: ')
        sys.stdout.buffer.write(result.stderr.encode('utf-8'))
    sys.stdout.buffer.write(b'\n')

    # Cleanup temp files
    try:
        os.remove('tests/_xlsx_generated.csv')
        os.remove('tests/_xlsx_validator.js')
    except OSError:
        pass

    sys.exit(result.returncode)


if __name__ == '__main__':
    main()
